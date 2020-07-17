#!/usr/bin/python3
import sys
import time
import socket
import os
import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from copy import copy
#from datetime import datetime

letter = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
#netData = '新余 2020/5/19-2020/5/25 100 30 50 20 1 2'
dataDir = r'd:\source\data'
#backupDir = r'd:\source\backup'

def insertToExcel(netData:str):
    netData = netData.split(' ')
    files = os.listdir(dataDir)
    os.chdir(dataDir)
    week = datetime.datetime.isoweekday(datetime.datetime.today())
    if week > 2:
        fileDate = datetime.datetime.today() + datetime.timedelta(2-week)
    else:
        fileDate = datetime.datetime.today() + datetime.timedelta(4+week)
    #fileDate = datetime.datetime.today()
    newFile = '掌上医保注册用户-{}.xlsx'.format(fileDate.strftime("%Y%m%d"))
    wb = load_workbook(filename = newFile)
    if netData[0] in wb.sheetnames: 
        sheetNow = wb['{}'.format(netData[0])]
        for i in range(1,10000):
            if sheetNow['A{}'.format(i)].value == None:
                lastDate = sheetNow['A{}'.format(i-1)].value
                n = 0
                for j in netData[1:]:
                    sheetNow['{}{}'.format(letter[n], i)].value = j
                    font1 = Font(size = 12, bold = True)
                    alignment1 = Alignment(horizontal = 'center', vertical = 'center')
                    sheetNow['{}{}'.format(letter[n], i)].font = font1
                    sheetNow['{}{}'.format(letter[n], i)].alignment = alignment1
                    sheetNow['{}{}'.format(letter[n], i)].number_format = '0'
                    sheetNow['A{}'.format(i)].number_format = 'yyyy/m/d\ h:mm;@'
                    #sheetNow['{}{}'.format(letter[n], i)] = copy(Font(sheetNow['{}{}'.format(letter[n], i - 1)]))
                    #print(Font(sheetNow['{}{}'.format(letter[n], i - 1)]))
                    #print(Alignment(sheetNow['{}{}'.format(letter[n], i - 1)]))
                    print(sheetNow['{}{}'.format(letter[n], i)].value, letter[n], i)
                    n += 1
                break
        wb.save(filename = 'filename.xlsx'.format(datetime.datetime.today().strftime("%Y%m%d")))
        os.chdir('..\\backup')
        wb.save(filename = 'filename.xlsx'.format(datetime.datetime.today().strftime("%Y%m%d")))
        print('保存成功')    
    else:
        print('没有对应的工作表')


if __name__ == '__main__':
    # 如果有参数，就接受第一个参数作为一个大字符串，切割后插入数据表中
    if len(sys.argv) == 2:
        insertToExcel(sys.argvp[1])
        exit(0)
    # 没有参数就启动网络监听
    sock = socket.socket()
    ip = "0.0.0.0"
    port = 9527
    sock.bind((ip,port))
    sock.listen()
    try:
        while True:
            conn, info = sock.accept()
            data = conn.recv(1024)
            dataDecode = data.decode()
            print(dataDecode)
            msg = '接收成功！ {}'.format(dataDecode)
            conn.send(msg.encode())
            if '!quit' in dataDecode:
                conn.close()
            insertToExcel(dataDecode.replace(' !quit\n', ''))
    finally:
        sock.close()

